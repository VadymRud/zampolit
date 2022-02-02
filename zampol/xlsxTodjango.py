import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "zampol.settings")
import django
django.setup()
from pprint import pprint
from datetime import datetime
from django.db import models
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from staff.models import Staff, SHPK
from osoba.models import ServiseID, MilitaryRank, Company
from django.utils.translation import gettext as _


file = 'D:\\Projects\\1pp.xlsx'
wb = load_workbook(filename=file)
ws = wb.worksheets[0]
last_column = len(list(ws.columns))
last_row = len(list(ws.rows))


my_list = {}
my_dict = []
for row in range(1, last_row + 1):
    my_dict = {}
    for column in range(2, last_column+1):
        column_letter = get_column_letter(column)

        if row > 1:
            my_dict[column] = ws['{}{}'.format(column_letter, str(row))].value
    my_list[row] = my_dict
# pprint(my_list)
i = 1
pidroz = Company()
for key, value in my_list.items():

    n2 = value.get(2)
    n3 = value.get(3)

    if n2 is not None and n3 is not None:
        if n2 in n3:
            print(value)
            pidroz = Company.objects.get_or_create(name=n2, unik=i)
            i += 1
        else:
            unicum = value.get(2)
            name_staff = value.get(3)
            shpk_name = value.get(4)
            vos = value.get(6)
            # staff = Staff.objects.get_or_create(unicum=unicum, name=name_staff, shpk=shpk, vos=vos)
            if not value.get(5):
                print('aaaaa')
                shpk = SHPK.objects.get_or_create(name=shpk_name)
                staff = Staff.objects.get_or_create(unicum=unicum, name=name_staff, shpk=shpk[0], vos=vos, salary=1,
                                                    tariff_category=1, company=pidroz[0])
            else:
                shpk = SHPK.objects.get_or_create(name=shpk_name)

                pib = value.get(7).split()
                mr = MilitaryRank.objects.get_or_create(name=value.get(5))
                osoba = ServiseID.objects.get_or_create(name=pib[1], sename=pib[0], third_name=pib[2],
                                                        military_ranks=mr[0], birth_date=datetime.now())
                staff = Staff.objects.get_or_create(unicum=unicum, name=name_staff, shpk=shpk[0], vos=vos, salary=1,
                                                    tariff_category=1, company=pidroz[0], ocoba=osoba[0])
                pprint(value)
    # for k, val in value.items():
    #     n3, n2 = '', ''
    #     n2 = value[2]
    #     n3 = value[3]
    #     if n2 == n3:
    #         print(value.items())

